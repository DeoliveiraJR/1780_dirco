# frontend/pages/upload.py
"""
Página de Upload de Dados Financeiros
"""
import streamlit as st
import pandas as pd
import numpy as np
import unicodedata
import re
from datetime import datetime
import json
import requests
import sys
import os

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from data_manager import set_dados_upload, get_dados_upload, salvar_upload_admin, eh_usuario_admin
from utils_ext.icons import render_page_header

# ==============================
# Configurações / Constantes
# ==============================
NUMERIC_COLS = [
    "CURVA_REALIZADO",
    "PROJETADO_ANALITICO",
    "PROJETADO_MERCADO",
    "PROJETADO_AJUSTADO",
]

MAP_MESES = {
    "janeiro": 1, "jan": 1, "jan.": 1,
    "fevereiro": 2, "fev": 2, "fev.": 2,
    "março": 3, "marco": 3, "mar": 3, "mar.": 3,
    "abril": 4, "abr": 4, "abr.": 4,
    "maio": 5, "mai": 5, "mai.": 5,
    "junho": 6, "jun": 6, "jun.": 6,
    "julho": 7, "jul": 7, "jul.": 7,
    "agosto": 8, "ago": 8, "ago.": 8,
    "setembro": 9, "set": 9, "set.": 9,
    "outubro": 10, "out": 10, "out.": 10,
    "novembro": 11, "nov": 11, "nov.": 11,
    "dezembro": 12, "dez": 12, "dez.": 12,
}

# ------------------- Normalização -------------------------------------------
def _norm_txt(s: str) -> str:
    if s is None:
        return ""
    s = unicodedata.normalize("NFKD", str(s))
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    return s.strip().lower()

def _norm_colname(c: str) -> str:
    """remove acentos, minúsculas e tira tudo que não for [a-z0-9]."""
    c = _norm_txt(c)
    return re.sub(r"[^a-z0-9]", "", c)

# Aliases: tudo mapeado para nomes canônicos (UPPER)
COL_ALIASES = {
    "datacompleta": "DATA_COMPLETA",
    "mes": "MES",
    "ano": "ANO",
    "codcategoria": "COD_CATEGORIA",
    "categoria": "CATEGORIA",
    "codproduto": "COD_PRODUTO",
    "produto": "PRODUTO",
    "curvarealizado": "CURVA_REALIZADO",
    "projetadoanalitico": "PROJETADO_ANALITICO",
    "projetadomercado": "PROJETADO_MERCADO",
    "projetadoajustado": "PROJETADO_AJUSTADO",

    # >>> Tipo de cliente: varias grafias comuns
    "tipocliente": "TIPO_CLIENTE",
    "clientetipo": "TIPO_CLIENTE",
    "tipodocliente": "TIPO_CLIENTE",
    "tpcliente": "TIPO_CLIENTE",
    "tpclient": "TIPO_CLIENTE",
    "tp_client": "TIPO_CLIENTE",     # normalizado vira 'tp_client' -> 'tpclient' sem underscore
}

def _rename_columns_flex(df: pd.DataFrame) -> pd.DataFrame:
    mapping = {}
    for c in df.columns:
        nc = _norm_colname(c)
        if nc in COL_ALIASES:
            mapping[c] = COL_ALIASES[nc]
    if mapping:
        df = df.rename(columns=mapping)
    return df

def _unificar_tipo_cliente(df: pd.DataFrame) -> pd.DataFrame:
    """
    Garante uma única coluna TIPO_CLIENTE:
    - Se existirem TIPO_CLIENTE e TP_CLIENTE (ou variantes não mapeadas), prioriza valor não vazio.
    - Remove colunas redundantes que restarem (ex.: TP_CLIENTE).
    """
    dff = df.copy()

    # 1) Se houver uma coluna remanescente 'TP_CLIENTE' (não mapeada por algum motivo),
    #    unifica no TIPO_CLIENTE.
    #    Obs.: após _rename_columns_flex, o comum é já ter TIPO_CLIENTE e NÃO ter TP_CLIENTE.
    possiveis_tp_cliente = [c for c in dff.columns if _norm_colname(c) in ("tpcliente", "tpclient", "tp_client")]
    has_tipo = "TIPO_CLIENTE" in dff.columns

    if not has_tipo and possiveis_tp_cliente:
        # Se não existe TIPO_CLIENTE, mas existe TP_CLIENTE -> cria TIPO_CLIENTE a partir dele
        fonte = possiveis_tp_cliente[0]
        dff["TIPO_CLIENTE"] = dff[fonte]
        has_tipo = True

    if has_tipo and possiveis_tp_cliente:
        # Se por acaso existirem as duas, unificar escolhendo valor não vazio
        fonte = possiveis_tp_cliente[0]
        # se TIPO_CLIENTE vazio/NaN, pega do TP_CLIENTE
        mask_vazia = dff["TIPO_CLIENTE"].isna() | (dff["TIPO_CLIENTE"].astype(str).str.strip() == "")
        dff.loc[mask_vazia, "TIPO_CLIENTE"] = dff.loc[mask_vazia, fonte]

    # Remover duplicatas/variantes se ainda estiverem presentes
    for c in possiveis_tp_cliente:
        if c != "TIPO_CLIENTE" and c in dff.columns:
            dff = dff.drop(columns=[c])

    return dff

# ------------------- Datas ---------------------------------------------------
def _parse_date_mixed(series: pd.Series) -> pd.Series:
    """
    Se numérico com "cara" de serial Excel -> origin=1899-12-30.
    Senão, parser textual com dayfirst=True.
    """
    s = series.copy()

    # Tenta serial Excel para valores numericos
    if np.issubdtype(s.dropna().infer_objects().dtype, np.number):
        vals = pd.to_numeric(s, errors="coerce")
        if (vals > 10000).sum() >= max(1, int(0.5 * vals.notna().sum())):
            try:
                return pd.to_datetime(vals, unit="D", origin="1899-12-30", errors="coerce")
            except Exception:
                pass

    return pd.to_datetime(s, errors="coerce", dayfirst=True)

# ------------------- Sanitização principal -----------------------------------
def _sanitize_df_for_system(df: pd.DataFrame) -> pd.DataFrame:
    dff = df.copy()

    # 1) Renomeia por aliases (inclui TP_CLIENTE -> TIPO_CLIENTE)
    dff = _rename_columns_flex(dff)

    # 2) Unifica colunas de cliente (se ainda restar TP_CLIENTE)
    dff = _unificar_tipo_cliente(dff)

    # 3) Garante colunas mínimas
    for col in ["MES", "ANO", "CATEGORIA", "PRODUTO", "DATA_COMPLETA"]:
        if col not in dff.columns:
            dff[col] = ""

    # 4) TIPO_CLIENTE opcional: se não existir ainda, define default
    if "TIPO_CLIENTE" not in dff.columns:
        dff["TIPO_CLIENTE"] = "NÃO INFORMADO"

    # 5) Auxiliares normalizados
    dff["CAT_N"]  = dff["CATEGORIA"].apply(_norm_txt)
    dff["PROD_N"] = dff["PRODUTO"].apply(_norm_txt)
    dff["MES_N"]  = dff["MES"].apply(_norm_txt)
    dff["CLI_N"]  = dff["TIPO_CLIENTE"].apply(_norm_txt)

    # 6) Datas coerentes
    dff["DATA_COMPLETA_DT"] = _parse_date_mixed(dff["DATA_COMPLETA"])

    # 7) Derivações ANO/MES quando faltarem
    dff["ANO_NUM"] = pd.to_numeric(dff["ANO"], errors="coerce")
    dff.loc[dff["ANO_NUM"].isna() & dff["DATA_COMPLETA_DT"].notna(), "ANO_NUM"] = dff["DATA_COMPLETA_DT"].dt.year
    dff["ANO_NUM"] = dff["ANO_NUM"].fillna(0).astype(int)

    dff["MES_NUM"] = dff["MES_N"].map(MAP_MESES)
    dff.loc[dff["MES_NUM"].isna() & dff["DATA_COMPLETA_DT"].notna(), "MES_NUM"] = dff["DATA_COMPLETA_DT"].dt.month
    dff["MES_NUM"] = dff["MES_NUM"].fillna(0).astype(int)

    # 8) Sanitização numérica robusta
    for col in NUMERIC_COLS:
        if col not in dff.columns:
            dff[col] = 0.0
        dff[col] = (
            dff[col]
            .replace(
                to_replace=[r"^\s*missing\s*$", r"^\s*n/?a\s*$", r"^\s*-\s*$", r"^\s*null\s*$", r"^\s*none\s*$"],
                value=np.nan,
                regex=True,
            )
        )
        dff[col] = pd.to_numeric(dff[col], errors="coerce").fillna(0.0).astype(float)

    return dff

def _df_to_json_records(df: pd.DataFrame) -> list[dict]:
    out = df.copy()
    for col in out.select_dtypes(include=["datetime64[ns]", "datetime64[ns, UTC]"]).columns:
        out[col] = out[col].dt.strftime("%Y-%m-%d")
    out = out.where(pd.notnull(out), None)
    return out.to_dict("records")

def _consolidar_duplicatas(df: pd.DataFrame, metodo: str = "sum") -> pd.DataFrame:
    chaves = ["ANO_NUM", "MES_NUM", "CAT_N", "PROD_N"]
    if "CLI_N" in df.columns:
        chaves.append("CLI_N")

    num_cols = [c for c in df.columns if c in NUMERIC_COLS]
    if metodo not in {"sum", "mean", "first"}:
        metodo = "sum"
    agg_num = {c: (metodo if metodo in {"sum", "mean"} else "first") for c in num_cols}
    outras = [c for c in df.columns if c not in chaves + num_cols]
    agg_out = {c: "first" for c in outras}

    return (
        df.sort_values(chaves)
          .groupby(chaves, as_index=False)
          .agg({**agg_num, **agg_out})
    )

# ==============================
# Página
# ==============================
def renderizar():
    # Header elegante padronizado
    render_page_header(
        "Upload de Dados",
        "fa-cloud-arrow-up",
        "Envie seus arquivos Excel com os dados de projeção"
    )
    
    # ============== CONTROLE DE PERMISSÕES ==============
    if not eh_usuario_admin():
        st.error("""
        🔒 **Acesso Restrito**
        
        Apenas usuários com permissão de **Administrador** podem fazer upload de dados.
        
        **O que você pode fazer:**
        - ✓ Visualizar a base de dados compartilhada
        - ✗ Fazer upload de novos dados (apenas admin)
        - ✓ Criar suas próprias simulações
        """)
        
        st.markdown("---")
        st.markdown("#### 📊 Base de Dados Atual")
        dados_carregados()
        return
    
    # ============== INTERFACE PARA ADMIN ==============
    st.info("""
    ✓ **Você é Administrador**
    
    Você pode fazer upload de novos arquivos que serão compartilhados com todos os usuários do sistema.
    """)
    
    tab1, tab2, tab3 = st.tabs(["📤 Carregar Dados", "📊 Dados Carregados", "📈 Índices Econômicos"])
    with tab1:
        upload_interface()
    with tab2:
        dados_carregados()
    with tab3:
        indices_carregados()

def upload_interface():
    col1, col2 = st.columns([2, 1])

    with col1:
        st.markdown("#### Selecione o arquivo Excel para importar")
        st.markdown("Aceita 1 ou 2 abas: **DADOS** (projeções) e/ou **INDICES_TESOU** (índices econômicos)")
        
        uploaded_file = st.file_uploader(
            "Escolha um arquivo Excel (.xlsx ou .xls)",
            type=["xlsx", "xls"],
            key="file_uploader",
        )
        if uploaded_file is not None:
            try:
                import openpyxl
                from io import BytesIO
                
                # Detectar abas disponíveis
                arquivo_io = BytesIO(uploaded_file.getvalue())
                wb = openpyxl.load_workbook(arquivo_io)
                abas = wb.sheetnames
                
                st.success(f"✅ Arquivo carregado com sucesso! Abas detectadas: {', '.join(abas)}")
                
                # Lê e mostra prévia de cada aba
                for aba_nome in abas:
                    if aba_nome.lower().strip() in ['dados', 'data']:
                        df_raw = pd.read_excel(uploaded_file, sheet_name=aba_nome)
                        st.info(f"📊 **Aba '{aba_nome}' (Projeções)**: {len(df_raw)} registros")
                        with st.expander(f"📋 Prévia de '{aba_nome}'"):
                            st.dataframe(df_raw.head(100), use_container_width=True)
                    
                    elif aba_nome.lower().strip() == 'indices_tesou':
                        df_indices = pd.read_excel(uploaded_file, sheet_name=aba_nome)
                        st.info(f"📈 **Aba '{aba_nome}' (Índices)**: {len(df_indices)} registros")
                        with st.expander(f"📋 Prévia de '{aba_nome}'"):
                            st.dataframe(df_indices.head(100), use_container_width=True)
                
                # Botão para confirmar
                if st.button("✔️ Confirmar e Carregar", type="primary", use_container_width=True):
                    processar_arquivo_completo(uploaded_file)
            except Exception as e:
                st.error(f"❌ Erro ao ler arquivo: {str(e)}")
                import traceback
                traceback.print_exc()

    with col2:
        st.markdown("#### Estrutura Esperada")
        st.markdown("""
**Aba DADOS - Colunas Obrigatórias:**

- DATA_COMPLETA (ex.: 15/01/2026)
- MES (ex.: janeiro | jan)
- ANO (ex.: 2026)
- COD_CATEGORIA
- CATEGORIA
- COD_PRODUTO
- PRODUTO
- CURVA_REALIZADO
- PROJETADO_ANALITICO
- PROJETADO_MERCADO
- PROJETADO_AJUSTADO

**Aba INDICES_TESOU (Opcional):**

Importada sem tratamento, apenas como está no arquivo
        """)
        if st.button("📥 Baixar Template", use_container_width=True):
            gerar_template()


def processar_arquivo_completo(uploaded_file):
    """Processa arquivo com múltiplas abas (dados + índices)"""
    try:
        import openpyxl
        from io import BytesIO
        
        arquivo_bytes = uploaded_file.getvalue()
        arquivo_io = BytesIO(arquivo_bytes)
        wb = openpyxl.load_workbook(arquivo_io)
        abas = wb.sheetnames
        
        st.info(f"📋 Abas detectadas no arquivo: {', '.join(abas)}")
        
        # Processa aba de DADOS
        aba_dados = next((a for a in abas if a.lower().strip() in ['dados', 'data']), None)
        if aba_dados:
            st.markdown(f"**ℹ️ Processando aba '{aba_dados}' (Projeções)...**")
            df_raw = pd.read_excel(uploaded_file, sheet_name=aba_dados)
            
            # Processa os dados de projeção (sem mostrar botão de salvar duplicado)
            processar_dados(df_raw, mostrar_botao_salvar=False)
            
            # Se admin, oferece ÚNICO botão para salvar na base compartilhada
            if eh_usuario_admin():
                st.markdown("---")
                st.markdown("#### 💾 Salvar na Base de Dados Compartilhada")
                st.info("✓ Como administrador, você pode salvar este arquivo completo (com ambas as abas) como a nova base de dados compartilhada para todos os usuários.")
                
                if st.button("💾 Confirmar e Carregar", type="primary", use_container_width=True, key="btn_salvar_completo"):
                    usuario_id = st.session_state.get("usuario_id", "")
                    if not usuario_id:
                        st.error("❌ Erro: usuario_id não definido no session_state")
                        st.info("Por favor, faça logout e login novamente")
                        return
                    
                    st.info(f"📤 Enviando arquivo para backend... (usuario_id: {usuario_id})")
                    
                    # Salva o arquivo Excel COMPLETO (com ambas as abas)
                    sucesso, mensagem = salvar_upload_admin(arquivo_bytes, uploaded_file.name, usuario_id)
                    
                    st.write(f"[DEBUG] Resultado: sucesso={sucesso}, msg='{mensagem}'")
                    
                    if sucesso:
                        st.success(f"✅ {mensagem}")
                        st.balloons()
                    else:
                        st.error(f"❌ {mensagem}")
                        st.info("💡 Dica: Verifique se o backend está rodando (python backend/run.py)")
        else:
            st.warning("⚠️ Nenhuma aba 'DADOS' ou 'DATA' encontrada no arquivo")
        
        # Verifica aba de INDICES_TESOU (apenas informa)
        aba_indices = next((a for a in abas if a.lower().strip() == 'indices_tesou'), None)
        if aba_indices:
            st.markdown(f"✅ **Aba '{aba_indices}' (Índices Econômicos)** detectada e será importada automaticamente!")
        
    except Exception as e:
        st.error(f"❌ Erro ao processar arquivo: {str(e)}")
        import traceback
        traceback.print_exc()


def processar_dados(df_raw: pd.DataFrame, mostrar_botao_salvar: bool = True):
    try:
        # Renomeia e unifica ANTES de validar required
        df_raw = _rename_columns_flex(df_raw)
        df_raw = _unificar_tipo_cliente(df_raw)

        required = [
            "DATA_COMPLETA", "MES", "ANO", "COD_CATEGORIA", "CATEGORIA",
            "COD_PRODUTO", "PRODUTO", "CURVA_REALIZADO",
            "PROJETADO_ANALITICO", "PROJETADO_MERCADO", "PROJETADO_AJUSTADO",
        ]
        if not all(c in df_raw.columns for c in required):
            st.error("❌ Arquivo com colunas incompletas. Verifique a estrutura esperada.")
            return

        df_clean = _sanitize_df_for_system(df_raw)

        chaves = ["ANO_NUM", "MES_NUM", "CAT_N", "PROD_N"]
        if "CLI_N" in df_clean.columns:
            chaves.append("CLI_N")
        if df_clean.duplicated(subset=chaves, keep=False).any():
            df_clean = _consolidar_duplicatas(df_clean, metodo="sum")

        df_clean = df_clean[(df_clean["MES_NUM"] >= 1) & (df_clean["MES_NUM"] <= 12)]
        df_clean = df_clean[df_clean["ANO_NUM"] > 0]
        df_clean = df_clean.drop_duplicates()
        
        # VALIDAÇÃO CRÍTICA: Garantir que há dados após limpeza
        if len(df_clean) == 0:
            st.error("""
            ❌ **Nenhum dado válido após processamento**
            
            Possíveis causas:
            - Meses não reconhecidos ou vazios
            - Anos não especificados ou inválidos
            - Valores numéricos ausentes nas colunas de projeção
            
            Por favor, verifique o arquivo e tente novamente.
            """)
            return

        set_dados_upload(df_clean)

        with st.expander("🔎 Visualizar dados LIMPOS (o que o sistema usará)"):
            st.dataframe(df_clean.head(200), use_container_width=True)

        dados_json = _df_to_json_records(df_clean)

        try:
            resp = requests.post(
                "http://localhost:5000/api/upload",
                json={"data": dados_json},
                timeout=10,
            )
            if resp.status_code == 200:
                st.session_state.dados_carregados = dados_json
                st.success("✅ Dados carregados no backend!")
                st.balloons()
            else:
                st.warning("⚠️ Backend respondeu com erro. Dados salvos localmente.")
                st.session_state.dados_carregados = dados_json
        except requests.exceptions.RequestException:
            st.warning("⚠️ Backend indisponível. Dados salvos localmente.")
            st.session_state.dados_carregados = dados_json
        
        # ============== BOTÃO DE SALVAR (apenas se mostrar_botao_salvar=True) ==============
        # Nota: Este botão é desabilitado quando chamado de processar_arquivo_completo
        # para evitar duplicação. O fluxo unificado é em processar_arquivo_completo()

    except Exception as e:
        st.error(f"❌ Erro ao processar dados: {str(e)}")

def dados_carregados():
    df = get_dados_upload()
    
    # Validar se há dados
    if df is None or len(df) == 0:
        st.info("ℹ️ Nenhum dado carregado ainda. Faça upload na aba anterior.")
        return
    
    # Validar colunas essenciais
    cols_requeridas = ["CATEGORIA", "PRODUTO", "MES_NUM"]
    colunas_faltantes = [c for c in cols_requeridas if c not in df.columns]
    
    if colunas_faltantes:
        st.error(f"""
        ❌ **Dados Incompletos**
        
        As colunas esperadas não foram encontradas: {', '.join(colunas_faltantes)}
        
        Por favor, verifique o arquivo e faça upload novamente.
        """)
        return

    st.markdown(f"#### Total de Registros (limpos): {len(df)}")

    c1, c2, c3 = st.columns(3)
    with c1:
        st.metric("Categorias", df["CATEGORIA"].nunique())
    with c2:
        st.metric("Produtos", df["PRODUTO"].nunique())
    with c3:
        st.metric("Períodos distintos", df["MES_NUM"].replace(0, np.nan).nunique())

    st.markdown("---")
    st.dataframe(df, use_container_width=True)

    if st.button("🗑️ Limpar Dados", use_container_width=True):
        st.session_state.dados_carregados = None
        set_dados_upload(pd.DataFrame())
        st.rerun()

def indices_carregados():
    """Visualiza os índices econômicos compartilhados"""
    
    try:
        # Calcular path do backend corretamente
        # __file__ = frontend/pages/upload.py
        # Precisamos chegar em 1780_dirco/backend
        current_file = os.path.abspath(__file__)  # .../frontend/pages/upload.py
        frontend_dir = os.path.dirname(os.path.dirname(current_file))  # .../frontend
        root_dir = os.path.dirname(frontend_dir)  # .../1780_dirco
        backend_dir = os.path.join(root_dir, 'backend')
        
        sys.path.insert(0, backend_dir)
        from database import carregar_indices_compartilhados, obter_metadados_ultimo_upload_indices, indices_existem
        
    except ImportError as e:
        st.error(f"❌ Erro ao importar funções de índices do backend: {str(e)}")
        return
    
    # Verifica se há índices importados
    if not indices_existem():
        st.info("ℹ️ Nenhum índice econômico foi importado ainda. Faça upload na aba 'Carregar Dados' com a aba INDICES_TESOU.")
        return
    
    # Carrega os índices
    df_indices = carregar_indices_compartilhados()
    
    if df_indices is None or len(df_indices) == 0:
        st.warning("⚠️ Erro ao carregar índices ou nenhum dado disponível")
        return
    
    # Mostra informações dos índices
    metadata = obter_metadados_ultimo_upload_indices()
    
    if metadata:
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("Total de Registros", len(df_indices))
        with col2:
            st.metric("Total de Colunas", len(df_indices.columns))
        with col3:
            data_upload = metadata.get("data_upload", "N/A")
            st.info(f"📅 Último upload: {data_upload[:10]}")
    
    st.markdown("---")
    
    # Lista as colunas disponíveis
    st.markdown("#### 📋 Colunas Disponíveis")
    st.markdown("**Campos principais:**")
    colunas_principais = ["DT_ALVO", "DT_PRJ", "VL_PJTD", "NM_IN"]
    for col in colunas_principais:
        if col in df_indices.columns:
            st.markdown(f"  - ✅ `{col}`")
        else:
            st.markdown(f"  - ⚠️ `{col}` (não encontrado)")
    
    st.markdown("---")
    st.markdown("#### 📊 Visualização dos Dados (Primeiros 100 registros)")
    st.dataframe(df_indices.head(100), use_container_width=True)
    
    # Estatísticas básicas para colunas numéricas
    st.markdown("---")
    st.markdown("#### 📈 Estatísticas Básicas")
    numeric_cols = df_indices.select_dtypes(include=['number']).columns
    if len(numeric_cols) > 0:
        st.dataframe(df_indices[numeric_cols].describe(), use_container_width=True)
    else:
        st.info("ℹ️ Nenhuma coluna numérica para estatísticas")

def gerar_template():
    meses = ['janeiro','fevereiro','março','abril','maio','junho',
             'julho','agosto','setembro','outubro','novembro','dezembro']

    categorias = [
        {'COD': 'CAT001', 'NOME': 'OPERACOES DE CREDITO - CARTEIRA AMPLIADA PAIS'},
        {'COD': 'CAT002', 'NOME': 'SERVICOS'},
        {'COD': 'CAT003', 'NOME': 'CAPTACOES'},
    ]
    produtos = [
        {'COD': 'PRD001', 'NOME': 'CREDITO PESSOAL'},
        {'COD': 'PRD002', 'NOME': 'EMPRESARIAL'},
        {'COD': 'PRD003', 'NOME': 'FUNDO X'},
    ]
    tipos_cliente = ['CLIENTE VAREJO', 'CLIENTE ATACADO', 'CLIENTE PRIVATE']

    dados = []
    for ano in [2025, 2026]:
        for mi, mes in enumerate(meses, start=1):
            for cat in categorias:
                for prod in produtos:
                    for cli in tipos_cliente:
                        dados.append({
                            'DATA_COMPLETA': f'15/{mi:02d}/{ano}',
                            'MES': mes,
                            'ANO': str(ano),
                            'COD_CATEGORIA': cat['COD'],
                            'CATEGORIA': cat['NOME'],
                            'COD_PRODUTO': prod['COD'],
                            'PRODUTO': prod['NOME'],
                            'TIPO_CLIENTE': cli,
                            'CURVA_REALIZADO': np.random.randint(100000, 1000000),
                            'PROJETADO_ANALITICO': np.random.randint(100000, 1000000),
                            'PROJETADO_MERCADO': np.random.randint(100000, 1000000),
                            'PROJETADO_AJUSTADO': np.random.randint(100000, 1000000),
                        })

    df_template = pd.DataFrame(dados)
    st.download_button(
        label="📥 Baixar Template (CSV)",
        data=df_template.to_csv(index=False).encode("utf-8"),
        file_name=f"template_dados_{datetime.now():%Y%m%d}.csv",
        mime="text/csv",
        use_container_width=True,
    )
